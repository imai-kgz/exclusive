import streamlit as st
import pandas as pd
import os
from io import BytesIO
from fpdf import FPDF
import openpyxl
import math
import requests
from datetime import datetime

# --- Settings ---
DATA_DIR = "uploaded_files"
os.makedirs(DATA_DIR, exist_ok=True)
FONT_REGULAR_PATH = "JetBrainsMono-Regular.ttf"
FONT_BOLD_PATH = "JetBrainsMono-Bold.ttf"
ITEMS_PER_PAGE = 30

# --- GitHub File URLs ---
GITHUB_FILE_URLS = [
    "https://raw.githubusercontent.com/imai-kgz/exclusive/main/Аутоимунка 2025.xlsx",
    "https://raw.githubusercontent.com/imai-kgz/exclusive/main/Прайс основной.xlsx"
]

# --- User Credentials (from st.secrets) ---
VALID_USERNAME = st.secrets.get("APP_USERNAME", "admin")
VALID_PASSWORD = st.secrets.get("APP_PASSWORD", "admin")

# --- Session State Initialization ---
def initialize_session_state():
    defaults = {
        'theme': 'dark', 'files': {}, 'selected_analyses': {},
        'price_edit_enabled': {}, 'show_selected': False,
        'logged_in': False, 'pages': {}
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# --- Core Functions ---

@st.cache_data
def load_data_from_excel(file_content):
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    sheet = wb.active
    data = []
    current_group = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=False), start=1):
        if len(row) < 5: continue
        title_cell, color0_cell, color1_cell, price_cell, time_cell = row[:5]
        title, price = title_cell.value, price_cell.value
        is_group_header = (price is None or price == '') and (title is not None and title != '')
        is_analysis_row = (price is not None and price != '') and (title is not None and title != '')
        if is_group_header: current_group = title
        elif is_analysis_row:
            try: price_value = float(price)
            except (ValueError, TypeError): price_value = 0.0
            color0_color = f"#{color0_cell.fill.fgColor.rgb[2:]}".lower() if color0_cell.fill.patternType == 'solid' and color0_cell.fill.fgColor.rgb and color0_cell.fill.fgColor.rgb not in ('00000000', 'FFFFFFFF') else None
            color1_color = f"#{color1_cell.fill.fgColor.rgb[2:]}".lower() if color1_cell.fill.patternType == 'solid' and color1_cell.fill.fgColor.rgb and color1_cell.fill.fgColor.rgb not in ('00000000', 'FFFFFFFF') else None
            data.append({
                'original_index': row_idx - 1, 'group': current_group, 'title': title, 'price': price_value,
                'color_0_text': str(color0_cell.value or ''), 'color_0_color': color0_color,
                'color_1_text': str(color1_cell.value or ''), 'color_1_color': color1_color,
                'time': time_cell.value or ''
            })
    return pd.DataFrame(data)

def save_changes_to_file(file_path, df_to_save):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        price_map = {row['title']: row['price'] for _, row in df_to_save.iterrows()}
        for row in sheet.iter_rows(min_row=1):
            if len(row) > 3 and row[0].value in price_map and (row[3].value is not None and row[3].value != ''):
                row[3].value = price_map[row[0].value]
        output = BytesIO()
        wb.save(output); output.seek(0)
        with open(file_path, 'wb') as f: f.write(output.getvalue())
        output.seek(0); return output.getvalue()
    except Exception as e:
        st.error(f"Ошибка при сохранении файла: {e}"); return None

@st.cache_resource
def load_files_from_github():
    files_data = {}
    for url in GITHUB_FILE_URLS:
        try:
            response = requests.get(url)
            response.raise_for_status()
            file_name = url.split('/')[-1]
            temp_path = os.path.join(DATA_DIR, file_name)
            with open(temp_path, "wb") as f: f.write(response.content)
            df = load_data_from_excel(response.content)
            files_data[file_name] = {'data': df, 'path': temp_path}
        except requests.exceptions.RequestException as e:
            st.error(f"Не удалось загрузить файл {url.split('/')[-1]}: {e}")
    return files_data

def handle_checkbox_change(file_name, idx):
    st.session_state.selected_analyses.setdefault(file_name, [])
    if idx in st.session_state.selected_analyses[file_name]:
        st.session_state.selected_analyses[file_name].remove(idx)
    else:
        st.session_state.selected_analyses[file_name].append(idx)

def generate_pdf_receipt(selected_analyses_data, custom_title=""):
    pdf = FPDF()
    pdf.add_page()
    try:
        pdf.add_font('DejaVu', '', FONT_REGULAR_PATH, uni=True)
        pdf.add_font('DejaVu', 'B', FONT_BOLD_PATH, uni=True)
    except Exception as e:
        st.error(f"Font Error: {e}. Ensure font files are in the same folder.")
        return None
    pdf_title = custom_title.strip() if custom_title.strip() else "Ваш Чек"
    pdf.set_font("DejaVu", size=16, style="B")
    pdf.cell(0, 10, txt=pdf_title, ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("DejaVu", size=11)
    total_sum_pdf, line_width = 0, 190
    for item in selected_analyses_data:
        color_texts = ' '.join(filter(None, [item.get('color_0_text'), item.get('color_1_text')])).strip()
        full_title = f"{item['title']} ({color_texts})" if color_texts else item['title']
        price, price_str = float(item['price']), f"{float(item['price']):,.2f} с"
        total_sum_pdf += price
        price_width, dot_width = pdf.get_string_width(price_str), pdf.get_string_width('.')
        lines, remaining_text = [], full_title
        while pdf.get_string_width(remaining_text) > 0:
            line = remaining_text
            while pdf.get_string_width(line) > line_width: line = line[:-1]
            if len(line) < len(remaining_text):
                split_pos = line.rfind(' ')
                if split_pos > 0: line = line[:split_pos]
            lines.append(line)
            remaining_text = remaining_text[len(line):].strip()
        if not lines: continue
        if len(lines) > 1:
            for line in lines[:-1]: pdf.cell(0, 8, txt=line, ln=True)
        last_line = lines[-1]
        last_line_width = pdf.get_string_width(last_line)
        if last_line_width + price_width < (line_width - 5):
            num_dots = int((line_width - last_line_width - price_width) / dot_width) if dot_width > 0 else 0
            dots = '.' * max(0, num_dots)
            pdf.cell(last_line_width, 8, txt=last_line, ln=False)
            pdf.cell(line_width - last_line_width - price_width, 8, txt=dots, ln=False)
            pdf.cell(price_width, 8, txt=price_str, ln=True, align="R")
        else:
            pdf.cell(0, 8, txt=last_line, ln=True)
            num_dots = int((line_width - price_width) / dot_width) if dot_width > 0 else 0
            dots = '.' * max(0, num_dots)
            pdf.cell(line_width - price_width, 8, txt=dots, ln=False)
            pdf.cell(price_width, 8, txt=price_str, ln=True, align="R")
    now = datetime.now()
    date_time_str = now.strftime("%d.%m.%Y %H:%M:%S")
    pdf.ln(5); pdf.set_font("DejaVu", size=10)
    pdf.cell(0, 8, txt=f"Дата формирования: {date_time_str}", ln=True, align="R")
    pdf.ln(2); pdf.set_font("DejaVu", size=14, style="B")
    pdf.cell(150, 10, txt="Итого:", ln=False, align="R")
    pdf.cell(40, 10, txt=f"{total_sum_pdf:,.2f} с", ln=True, align="R")
    return pdf.output(dest="S").encode('latin-1')

@st.cache_data
def get_pdf_bytes_cached(_selected_items_tuple, title):
    rehydrated_list = [dict(item) for item in _selected_items_tuple]
    return generate_pdf_receipt(rehydrated_list, title)

def get_text_color(bg_color):
    if not bg_color or len(bg_color) != 7: return '#000000'
    r, g, b = int(bg_color[1:3], 16), int(bg_color[3:5], 16), int(bg_color[5:7], 16)
    return '#000000' if (0.299 * r + 0.587 * g + 0.114 * b) / 255 > 0.5 else '#ffffff'

def get_color_display_html(item, col_prefix):
    text, color = item.get(col_prefix + '_text', ''), item.get(col_prefix + '_color')
    if text:
        text_color = get_text_color(color)
        style = f"background-color:{color}; color:{text_color}; padding:2px 5px; border-radius:3px; font-size:14px;" if color else ""
        return f"<span style='{style}'>{text}</span>"
    return f"<div style='background-color:{color}; height:20px; width:100%; border-radius:3px;'></div>" if color else ""

def get_theme_styles(theme='dark'):
    common_styles = """<style> #search-container { position: fixed; top: 0; left: 0; width: 100%; padding: 10px 20px; z-index: 1001; box-shadow: 0 2px 4px rgba(0,0,0,0.2); } .main-content { margin-top: 80px; } .price-text { text-align: left; font-weight: bold; } </style>"""
    if theme == 'light': return common_styles + """<style> .stApp { background-color: #FFFFFF; color: #111111; font-size: 17px; } .stExpander, .stNumberInput input { background-color: #EEEEEE; } .stButton > button { background-color: #4CAF50; color: white; } p, label, .stMarkdown, h1, h2, h3 { color: #111111 !important; } #search-container { background-color: #FFFFFF; } .price-text { color: #000 !important; } </style>"""
    return common_styles + """<style> .stApp { background-color: #0E1117; color: #FAFAFA; font-size: 17px; } .stExpander, .stNumberInput input { background-color: #262730; } .stButton > button { background-color: #3f51b5; color: white; } p, label, .stMarkdown, h1, h2, h3 { color: #FAFAFA !important; } #search-container { background-color: #0E1117; } .price-text { color: #fff !important; } </style>"""

def login_form():
    st.set_page_config(page_title="Авторизация", layout="centered")
    st.markdown(get_theme_styles(st.session_state.theme), unsafe_allow_html=True)
    with st.container(border=True):
        st.title("Авторизация")
        username = st.text_input("Имя пользователя", key="username")
        password = st.text_input("Пароль", type="password", key="password")
        if st.button("Войти", type="primary"):
            if username == VALID_USERNAME and password == VALID_PASSWORD:
                st.session_state.logged_in = True
                st.rerun()
            else: st.error("Неверное имя пользователя или пароль.")

def display_sidebar():
    with st.sidebar:
        st.header("Настройки")
        is_dark = st.toggle("Темная тема", value=(st.session_state.theme == 'dark'))
        if is_dark != (st.session_state.theme == 'dark'):
            st.session_state.theme = 'dark' if is_dark else 'light'; st.rerun()
        st.markdown("---")
        st.header("Управление сессией")
        if st.button("Выйти"):
            for key in list(st.session_state.keys()): del st.session_state[key]
            st.rerun()

def clear_all_selections():
    st.session_state.selected_analyses.clear()
    if "check_title_input" in st.session_state:
        st.session_state.check_title_input = ""

def force_rerun_callback():
    pass

def main_app():
    st.set_page_config(page_title="Калькулятор анализов", layout="wide")
    st.markdown(get_theme_styles(st.session_state.theme), unsafe_allow_html=True)
    
    display_sidebar()

    if 'files' not in st.session_state or not st.session_state.files:
        with st.spinner("Загрузка и обработка файлов из GitHub..."):
            st.session_state.files = load_files_from_github()
        if not st.session_state.files:
            st.error("Не удалось загрузить основные файлы из GitHub."); return
        st.rerun()

    st.markdown('<div class="main-content">', unsafe_allow_html=True)
    st.title("Калькулятор анализов")
    st.markdown('<div id="search-container">', unsafe_allow_html=True)
    search_term = st.text_input("Поиск", placeholder="🔍 Поиск по названию, группе или тексту", label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

    tab_names = list(st.session_state.files.keys())
    tabs = st.tabs(tab_names)

    for i, file_name in enumerate(tab_names):
        with tabs[i]:
            if file_name not in st.session_state.files: continue
            file_info = st.session_state.files[file_name]
            df = file_info['data']
            df['group'] = df['group'].fillna('')
            manage_cols = st.columns([2, 2, 1])
            is_edit_mode = manage_cols[0].toggle("Редактировать цены", key=f"edit_{file_name}")
            if is_edit_mode and manage_cols[1].button("💾 Сохранить", key=f"save_{file_name}"):
                if save_changes_to_file(file_info['path'], df):
                    st.success("Сохранено!"); st.cache_resource.clear(); st.rerun()
                else: st.error("Ошибка сохранения.")
            filtered_df = df
            if search_term:
                search_lower = search_term.lower()
                item_mask = (df['title'].str.lower().str.contains(search_lower, na=False) | df['color_0_text'].str.lower().str.contains(search_lower, na=False) | df['color_1_text'].str.lower().str.contains(search_lower, na=False))
                matching_groups = df[df['group'].str.lower().str.contains(search_lower, na=False)]['group'].unique()
                group_mask = df['group'].isin(matching_groups)
                filtered_df = df[item_mask | group_mask]
            if filtered_df.empty: st.info("Ничего не найдено.")
            else:
                for group_name, group_df in sorted(filtered_df.groupby('group')):
                    if not group_name: continue
                    with st.expander(group_name, expanded=bool(search_term)):
                        total_items, total_pages = len(group_df), math.ceil(len(group_df) / ITEMS_PER_PAGE)
                        page_key = f"page_{file_name}_{group_name}"
                        st.session_state.pages.setdefault(page_key, 1)
                        current_page = st.session_state.pages[page_key]
                        start_idx, end_idx = (current_page - 1) * ITEMS_PER_PAGE, current_page * ITEMS_PER_PAGE
                        paginated_df = group_df.iloc[start_idx:end_idx]
                        for idx, row in paginated_df.iterrows():
                            cols = st.columns([5, 1.5, 1.5, 2])
                            is_checked = idx in st.session_state.selected_analyses.get(file_name, [])
                            cols[0].checkbox(row['title'], value=is_checked, key=f"check_{file_name}_{idx}", on_change=handle_checkbox_change, args=(file_name, idx))
                            cols[1].markdown(get_color_display_html(row, 'color_0'), unsafe_allow_html=True)
                            cols[2].markdown(get_color_display_html(row, 'color_1'), unsafe_allow_html=True)
                            with cols[3]:
                                if is_edit_mode:
                                    new_price = st.number_input("Цена", value=float(row['price']), key=f"price_{file_name}_{idx}", label_visibility="collapsed")
                                    if new_price != row['price']: st.session_state.files[file_name]['data'].at[idx, 'price'] = new_price
                                else:
                                    price_str = f"{int(row['price'])}" if row['price'] == int(row['price']) else f"{row['price']:.2f}"
                                    st.markdown(f"<p class='price-text'>{price_str} сом</p>", unsafe_allow_html=True)
                        if total_pages > 1:
                            st.markdown("---")
                            nav_cols = st.columns([1, 1, 1])
                            if nav_cols[0].button("< Пред.", key=f"prev_{page_key}", disabled=(current_page == 1)):
                                st.session_state.pages[page_key] -= 1; st.rerun()
                            nav_cols[1].markdown(f"<div style='text-align: center;'>Стр {current_page}/{total_pages}</div>", unsafe_allow_html=True)
                            if nav_cols[2].button("След. >", key=f"next_{page_key}", disabled=(current_page == total_pages)):
                                st.session_state.pages[page_key] += 1; st.rerun()

    st.markdown("---")
    
    total_sum = 0.0
    selected_data = []
    for fname, indices in st.session_state.selected_analyses.items():
        if indices and fname in st.session_state.files:
            df_file = st.session_state.files[fname]['data']
            valid_indices = df_file.index.intersection(indices)
            total_sum += df_file.loc[valid_indices, 'price'].sum()
            for _, row in df_file.loc[valid_indices].iterrows(): selected_data.append(row.to_dict())
    
    if selected_data:
        user_title = st.text_input(
            "Имя пациента или название чека (для PDF)",
            placeholder="Например: Ниязбек уул Эркинбек",
            key="check_title_input",
            on_change=force_rerun_callback
        )

        bottom_cols = st.columns([1, 1])
        
        now = datetime.now()
        date_str = now.strftime(" %Y-%m-%d")
        clean_user_title = user_title.strip()
        file_base_name = clean_user_title if len(clean_user_title) > 0 else f"Чек_{date_str}"
        pdf_file_name = f"{file_base_name + date_str}.pdf"
        print(clean_user_title)
        frozen_selected_data = tuple(frozenset(item.items()) for item in selected_data)
        pdf_bytes = get_pdf_bytes_cached(frozen_selected_data, clean_user_title)
        
        st.markdown("---")


        button_label = f"🧾 Выбрано: {len(selected_data)} поз. на сумму {total_sum:,.2f} сом"
        if st.button(button_label, key="toggle_selected"):
            st.session_state.show_selected = not st.session_state.show_selected
        
        if st.session_state.show_selected:
            st.markdown("#### Выбранные анализы:")
            for item in selected_data:
                cols = st.columns([5, 1.5, 1.5, 2])
                cols[0].write(f"- {item['title']}")
                cols[1].markdown(get_color_display_html(item, 'color_0'), unsafe_allow_html=True)
                cols[2].markdown(get_color_display_html(item, 'color_1'), unsafe_allow_html=True)
                cols[3].markdown(f"<p class='price-text' style='text-align: right;'>{item['price']:,.2f} сом</p>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        if pdf_bytes:
            bottom_cols[0].download_button(
                "Скачать чек (PDF)", data=pdf_bytes, file_name=pdf_file_name, 
                mime="application/pdf", key="download_pdf_button"
            )
        
        bottom_cols[1].button("🗑️ Очистить все", key="clear_all_button", on_click=clear_all_selections)

    st.markdown(f"""<div style="position: fixed; bottom: 10px; right: 100px; background-color: #3f51b5; padding: 15px; border-radius: 10px; box-shadow: 2px 2px 5px rgba(0,0,0,0.2);"><h3 style="margin: 0; color: #ffffff;">Итого: {total_sum:,.2f} сом</h3></div>""", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# --- App Entry Point ---
initialize_session_state()
if not st.session_state.logged_in:
    login_form()
else:
    main_app()
